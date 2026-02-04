import argparse
import os
import re
from typing import Iterable, Optional

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

def clean_text(value: Optional[str]) -> Optional[str]:
    """
    Normalize editorial bracket markup in text cells.

    Behavior:
    - Replace '[- ... ]' spans with inner content (e.g., '[-나의]' -> '나의', '[-(792)]' -> '(792)')
    - Clean residual artifacts like '],[-' and '][-'
    - Do NOT modify general whitespace or punctuation spacing (handled elsewhere)
    """
    if value is None:
        return value
    if not isinstance(value, str):
        return value

    s = value

    # Replace editorial bracket spans: '[- ... ]' -> inner content
    # Captures everything up to the closing bracket (excluding newlines)
    s = re.sub(r"\[-\s*([^\]]+)\]", r"\1", s)

    # Clean residual artifacts created by segmentation:
    # ],[-   -> comma + space
    s = re.sub(r"\],\s*\[-", ", ", s)
    # ][-    -> single space
    s = re.sub(r"\]\s*\[-", " ", s)
    # Lone closing brackets ']' left behind (rare): remove
    s = s.replace("]", "")
    # Lone open markers '[-' (rare): remove
    s = s.replace("[-", "")

    # 그대로 반환 (공백/구두점 정규화는 다른 단계에서 처리됨)
    return s

def iter_xlsx_files(src_dir: str) -> Iterable[str]:
    for root, _, files in os.walk(src_dir):
        for f in files:
            if f.lower().endswith(".xlsx"):
                yield os.path.join(root, f)

def clean_workbook(
    path: str,
    out_path: Optional[str] = None,
    columns_to_clean: Optional[Iterable[str]] = None,
) -> str:
    """
    Clean bracket markup in the given workbook.

    - If columns_to_clean is provided, only clean those column headers (exact match).
      Otherwise, clean all string cells.
    - Writes to out_path if provided; otherwise overwrites the original.
    Returns the path written.
    """
    wb: Workbook = load_workbook(filename=path)
    for ws in wb.worksheets:
        header_map = None
        if columns_to_clean:
            # Try to build a header map from the first row
            header_map = {}
            first_row = (
                next(ws.iter_rows(min_row=1, max_row=1)) if ws.max_row >= 1 else []
            )
            for idx, cell in enumerate(first_row, start=1):
                header = str(cell.value) if cell.value is not None else None
                if header:
                    header_map[header] = idx

        if header_map and columns_to_clean:
            target_cols = [header_map.get(col) for col in columns_to_clean]
            target_cols = [c for c in target_cols if c is not None]
            if target_cols:
                for row in ws.iter_rows(min_row=2):
                    for cidx in target_cols:
                        cell = row[cidx - 1]
                        cell.value = clean_text(cell.value)
                continue  # next sheet

        # Fallback: clean all string cells
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    cell.value = clean_text(cell.value)

    write_path = out_path or path
    os.makedirs(os.path.dirname(write_path), exist_ok=True)
    wb.save(write_path)
    return write_path

def main():
    parser = argparse.ArgumentParser(
        description="Strip editorial bracket markup ([-...]) from xlsx files."
    )
    parser.add_argument(
        "--src-dir",
        type=str,
        default="xlsx",
        help="Source directory containing .xlsx files (recursively processed). Default: xlsx",
    )
    parser.add_argument(
        "--dst-dir",
        type=str,
        default=None,
        help="Destination root directory for cleaned files. If omitted, files are overwritten in place.",
    )
    parser.add_argument(
        "--columns",
        type=str,
        default=None,
        help='Comma-separated column headers to clean (e.g., "원문,번역문"). If omitted, cleans all string cells.',
    )

    args = parser.parse_args()

    columns = [c.strip() for c in args.columns.split(",")] if args.columns else None
    for xlsx_path in iter_xlsx_files(args.src_dir):
        rel = os.path.relpath(xlsx_path, args.src_dir)
        if args.dst_dir:
            out_path = os.path.join(args.dst_dir, rel)
        else:
            out_path = None  # in-place

        written = clean_workbook(
            path=xlsx_path,
            out_path=out_path,
            columns_to_clean=columns,
        )
        print(f"Cleaned: {xlsx_path} -> {written}")

if __name__ == "__main__":
    main()
