"""
사서삼경을 제외한 모든 책의 문단병렬/문장병렬/구병렬 xlsx 파일을 합치는 스크립트.
출력: dataset/paragraph.xlsx, dataset/sentence.xlsx, dataset/phrase.xlsx
"""

import os
import sys
from pathlib import Path
import openpyxl
import logging

logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger(__name__)

BASE_DIR = Path(__file__).parent
XLSX_DIR = BASE_DIR / 'xlsx'
OUTPUT_DIR = BASE_DIR / 'dataset'

# 제외할 폴더
EXCLUDE_DIRS = {'사서삼경'}

# 병렬 유형 매핑
PARALLEL_TYPES = {
    '문단병렬': 'paragraph',
    '문장병렬': 'sentence',
    '구병렬': 'phrase',
}

# 각 유형별 헤더
HEADERS = {
    '문단병렬': ['책명', '문단식별자', '원문', '번역문'],
    '문장병렬': ['책명', '문단식별자', '문장식별자', '원문', '번역문'],
    '구병렬': ['책명', '문장식별자', '구식별자', '원문', '번역문'],
}


def collect_xlsx_files():
    """사서삼경을 제외한 모든 병렬 xlsx 파일을 수집."""
    files = {'문단병렬': [], '문장병렬': [], '구병렬': []}

    assert XLSX_DIR.exists(), f"xlsx 폴더를 찾을 수 없습니다: {XLSX_DIR}"

    for item in sorted(XLSX_DIR.iterdir()):
        if item.is_dir():
            if item.name in EXCLUDE_DIRS:
                logger.info(f"제외: {item.name}")
                continue

            # 폴더 내 xlsx 파일 수집
            for xlsx_file in sorted(item.glob('*.xlsx')):
                for ptype in PARALLEL_TYPES:
                    if xlsx_file.name.endswith(f'_{ptype}.xlsx'):
                        book_name = xlsx_file.name.replace(f'_{ptype}.xlsx', '')
                        files[ptype].append((book_name, xlsx_file))
                        break

        elif item.is_file() and item.suffix == '.xlsx':
            # 루트 레벨 xlsx 파일 (사정전훈의자치통감강목 등)
            for ptype in PARALLEL_TYPES:
                if item.name.endswith(f'_{ptype}.xlsx'):
                    book_name = item.name.replace(f'_{ptype}.xlsx', '')
                    files[ptype].append((book_name, item))
                    break

    return files


def read_xlsx_data(file_path):
    """xlsx 파일에서 데이터를 읽어 리스트로 반환 (헤더 제외)."""
    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb.active
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:  # 헤더 스킵
            continue
        rows.append(list(row))
    wb.close()
    return rows


def merge_and_save(ptype, file_list, output_path):
    """병렬 유형별 파일을 합쳐서 저장."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = PARALLEL_TYPES[ptype]

    # 헤더 작성
    ws.append(HEADERS[ptype])

    total_rows = 0
    for book_name, file_path in file_list:
        rows = read_xlsx_data(file_path)
        for row in rows:
            ws.append([book_name] + row)
        total_rows += len(rows)
        logger.info(f"  {book_name}: {len(rows)}행")

    wb.save(output_path)
    wb.close()
    logger.info(f"=> {output_path.name}: 총 {total_rows}행 저장 완료\n")
    return total_rows


def main():
    sys.stdout.reconfigure(encoding='utf-8')

    # 출력 폴더 생성
    OUTPUT_DIR.mkdir(exist_ok=True)

    # 파일 수집
    files = collect_xlsx_files()

    for ptype in PARALLEL_TYPES:
        logger.info(f"\n{'='*50}")
        logger.info(f"{ptype} ({PARALLEL_TYPES[ptype]}) - {len(files[ptype])}개 파일")
        logger.info(f"{'='*50}")

        if not files[ptype]:
            logger.warning(f"  {ptype} 파일이 없습니다.")
            continue

        output_path = OUTPUT_DIR / f"{PARALLEL_TYPES[ptype]}.xlsx"
        merge_and_save(ptype, files[ptype], output_path)

    logger.info("모든 병합 완료!")


if __name__ == '__main__':
    main()
